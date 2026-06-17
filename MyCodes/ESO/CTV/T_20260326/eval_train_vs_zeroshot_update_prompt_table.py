#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Evaluate train vs zero-shot predictions on test set and update prompt table.

What this script does:
1) Uses ONLY patients present in train prediction folder (expected 28 cases).
2) Strictly matches patient IDs by numeric part:
   - GT: test_nii/p_xx/CTV.nii.gz
   - Pred: best_mask/CTV_0xx.nii.gz
3) Computes:
   - Dice_3D (train / zero-shot)
   - HD95_3D (train / zero-shot)
4) Updates FIRST sheet in prompt_layer_merge_train_zeroshot.xlsx:
   - Keeps Patient_ID / Lower_Bound_ID / Upper_Bound_ID
   - Overwrites Dice3D values with current computation
   - Adds HD95_3D with same two-subcolumn format (zero-shot/train)
   - All metric values to 2 decimals
5) Creates/replaces SECOND sheet with per-slice metrics on GT-positive slices:
   - Patient_ID, Slice_ID, Lower_Bound_ID, Upper_Bound_ID
   - Dice2D (zero-shot/train), HD95_2D (zero-shot/train)
   - Values to 2 decimals
6) Checks old Dice values in first sheet; reports mismatched cases.
"""

from __future__ import annotations

import argparse
import math
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import SimpleITK as sitk
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


@dataclass
class CasePaths:
    patient_id: str
    gt_path: Path
    train_pred_path: Path
    zero_pred_path: Path
    nnunet_pred_path: Path


def parse_patient_id_from_gt_folder(name: str) -> Optional[str]:
    m = re.fullmatch(r"p_(\d+)", name)
    if not m:
        return None
    return f"CTV_{int(m.group(1)):03d}"


def parse_patient_id_from_pred_file(name: str) -> Optional[str]:
    m = re.fullmatch(r"CTV_(\d+)\.nii(\.gz)?", name)
    if not m:
        return None
    return f"CTV_{int(m.group(1)):03d}"


def read_nii(path: Path) -> Tuple[np.ndarray, sitk.Image]:
    img = sitk.ReadImage(str(path))
    arr = sitk.GetArrayFromImage(img)  # [Z, H, W]
    return arr, img


def to_binary(arr: np.ndarray) -> np.ndarray:
    return (arr > 0).astype(np.uint8)


def dice_nd(pred: np.ndarray, gt: np.ndarray, eps: float = 1e-8) -> float:
    pred_b = pred.astype(bool)
    gt_b = gt.astype(bool)
    inter = np.logical_and(pred_b, gt_b).sum()
    denom = pred_b.sum() + gt_b.sum()
    if denom == 0:
        return 1.0
    return float((2.0 * inter + eps) / (denom + eps))


def hd95_binary(pred: np.ndarray, gt: np.ndarray, spacing: Tuple[float, ...]) -> float:
    """
    Symmetric HD95 for 2D or 3D binary masks using SimpleITK distance maps.
    Returns NaN if either mask is empty.
    """
    pred = to_binary(pred)
    gt = to_binary(gt)
    if pred.sum() == 0 or gt.sum() == 0:
        return float("nan")

    pred_img = sitk.GetImageFromArray(pred)
    gt_img = sitk.GetImageFromArray(gt)
    pred_img.SetSpacing(tuple(float(x) for x in spacing))
    gt_img.SetSpacing(tuple(float(x) for x in spacing))

    pred_surface = sitk.LabelContour(pred_img)
    gt_surface = sitk.LabelContour(gt_img)

    dm_gt = sitk.Abs(
        sitk.SignedMaurerDistanceMap(
            gt_img, squaredDistance=False, useImageSpacing=True
        )
    )
    dm_pred = sitk.Abs(
        sitk.SignedMaurerDistanceMap(
            pred_img, squaredDistance=False, useImageSpacing=True
        )
    )

    pred_surface_arr = sitk.GetArrayViewFromImage(pred_surface) > 0
    gt_surface_arr = sitk.GetArrayViewFromImage(gt_surface) > 0
    dm_gt_arr = sitk.GetArrayViewFromImage(dm_gt)
    dm_pred_arr = sitk.GetArrayViewFromImage(dm_pred)

    d_pred_to_gt = dm_gt_arr[pred_surface_arr]
    d_gt_to_pred = dm_pred_arr[gt_surface_arr]

    all_d = np.concatenate([d_pred_to_gt, d_gt_to_pred]).astype(np.float64)
    if all_d.size == 0:
        return float("nan")
    return float(np.percentile(all_d, 95))


def get_gt_bounds(gt_zyx: np.ndarray) -> Tuple[int, int, List[int]]:
    pos = np.where(gt_zyx.reshape(gt_zyx.shape[0], -1).any(axis=1))[0]
    if len(pos) == 0:
        return -1, -1, []
    return int(pos.min()), int(pos.max()), [int(x) for x in pos.tolist()]


def collect_case_paths(
    gt_root: Path,
    train_pred_root: Path,
    zero_pred_root: Path,
    nnunet_pred_root: Path,
) -> List[CasePaths]:
    gt_map: Dict[str, Path] = {}
    for p in sorted(gt_root.iterdir()):
        if not p.is_dir():
            continue
        pid = parse_patient_id_from_gt_folder(p.name)
        if pid is None:
            continue
        gt_file = p / "CTV.nii.gz"
        if gt_file.exists():
            gt_map[pid] = gt_file

    train_map: Dict[str, Path] = {}
    for f in sorted(train_pred_root.glob("*.nii*")):
        pid = parse_patient_id_from_pred_file(f.name)
        if pid is None:
            continue
        train_map[pid] = f

    zero_map: Dict[str, Path] = {}
    for f in sorted(zero_pred_root.glob("*.nii*")):
        pid = parse_patient_id_from_pred_file(f.name)
        if pid is None:
            continue
        zero_map[pid] = f

    nnunet_map: Dict[str, Path] = {}
    for f in sorted(nnunet_pred_root.glob("*.nii*")):
        pid = parse_patient_id_from_pred_file(f.name)
        if pid is None:
            continue
        nnunet_map[pid] = f

    # Only evaluate cases present in train predictions.
    case_ids = sorted(
        train_map.keys(),
        key=lambda x: int(re.search(r"(\d+)$", x).group(1)) if re.search(r"(\d+)$", x) else 10**9,
    )
    cases: List[CasePaths] = []
    missing = []
    for pid in case_ids:
        g = gt_map.get(pid)
        z = zero_map.get(pid)
        t = train_map.get(pid)
        n = nnunet_map.get(pid)
        if g is None or z is None or t is None or n is None:
            missing.append(pid)
            continue
        cases.append(
            CasePaths(
                patient_id=pid,
                gt_path=g,
                train_pred_path=t,
                zero_pred_path=z,
                nnunet_pred_path=n,
            )
        )

    if missing:
        print(f"[WARN] Missing GT/zero/train for {len(missing)} cases, skipped: {missing}")

    print(f"[INFO] Cases selected from train predictions: {len(case_ids)}")
    print(f"[INFO] Cases evaluated: {len(cases)}")
    return cases


def cell_float(v) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s == "":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def setup_main_sheet_headers(ws):
    # Re-run safety: clear existing merged ranges first.
    if ws.merged_cells.ranges:
        for rng in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(rng))

    # A: Patient_ID, B: Lower_Bound_ID, C: Upper_Bound_ID
    # D:F Dice3D (zero-shot/train/nnunet_crop), G:I HD95_3D (zero-shot/train/nnunet_crop)
    ws["A1"] = "Patient_ID"
    ws["B1"] = "Lower_Bound_ID"
    ws["C1"] = "Upper_Bound_ID"
    ws["D1"] = "Dice3D"
    ws["G1"] = "HD95_3D"

    ws["D2"] = "SAM2_zeroshot"
    ws["E2"] = "SAM2_train"
    ws["F2"] = "nnunet_crop"
    ws["G2"] = "SAM2_zeroshot"
    ws["H2"] = "SAM2_train"
    ws["I2"] = "nnunet_crop"

    # Merge fixed columns over two rows
    for rng in ["A1:A2", "B1:B2", "C1:C2", "D1:F1", "G1:I1"]:
        ws.merge_cells(rng)

    center = Alignment(horizontal="center", vertical="center")
    bold = Font(bold=True)
    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=9):
        for cell in row:
            cell.alignment = center
            cell.font = bold


def write_main_sheet(
    ws,
    results_3d: Dict[str, Dict[str, float]],
    old_dice_mismatch: List[Tuple[str, Optional[float], Optional[float], Optional[float], Optional[float]]],
):
    setup_main_sheet_headers(ws)

    # Build sorted patient list from computed results.
    pids = sorted(
        results_3d.keys(),
        key=lambda x: int(re.search(r"(\d+)$", x).group(1)) if re.search(r"(\d+)$", x) else 10**9,
    )

    # Clear old data rows (row>=3, first 9 cols)
    if ws.max_row >= 3:
        for r in range(3, ws.max_row + 1):
            for c in range(1, 10):
                ws.cell(r, c).value = None

    row = 3
    for pid in pids:
        d = results_3d[pid]
        ws.cell(row, 1, pid)
        ws.cell(row, 2, d["lower"])
        ws.cell(row, 3, d["upper"])
        ws.cell(row, 4, d["dice_zero"])
        ws.cell(row, 5, d["dice_train"])
        ws.cell(row, 6, d["dice_nnunet"])
        ws.cell(row, 7, d["hd95_zero"])
        ws.cell(row, 8, d["hd95_train"])
        ws.cell(row, 9, d["hd95_nnunet"])
        row += 1

    center = Alignment(horizontal="center", vertical="center")
    for r in range(3, row):
        for c in range(1, 10):
            ws.cell(r, c).alignment = center

    # Number format to 2 decimals for metrics
    for r in range(3, row):
        for c in [4, 5, 6, 7, 8, 9]:
            ws.cell(r, c).number_format = "0.00"

    # Auto width
    for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I"]:
        ws.column_dimensions[col].width = 16

    if old_dice_mismatch:
        print(f"[WARN] Dice mismatch found in {len(old_dice_mismatch)} cases (old vs recomputed):")
        for pid, old_zero, new_zero, old_train, new_train in old_dice_mismatch:
            print(
                f"  - {pid}: zero old={old_zero} new={new_zero:.4f}, "
                f"train old={old_train} new={new_train:.4f}"
            )
    else:
        print("[INFO] No Dice mismatch against old table values (within tolerance).")


def write_slice_sheet(wb, per_slice_rows: List[Dict]):
    sheet_name = "Per_Slice_Metrics"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # Header structure:
    # A Patient_ID, B Slice_ID, C Lower_Bound_ID, D Upper_Bound_ID,
    # E:G Dice2D zero/train/nnunet, H:J HD95_2D zero/train/nnunet
    ws["A1"] = "Patient_ID"
    ws["B1"] = "Slice_ID"
    ws["C1"] = "Lower_Bound_ID"
    ws["D1"] = "Upper_Bound_ID"
    ws["E1"] = "Dice2D"
    ws["H1"] = "HD95_2D"

    ws["E2"] = "SAM2_zeroshot"
    ws["F2"] = "SAM2_train"
    ws["G2"] = "nnunet_crop"
    ws["H2"] = "SAM2_zeroshot"
    ws["I2"] = "SAM2_train"
    ws["J2"] = "nnunet_crop"

    for rng in ["A1:A2", "B1:B2", "C1:C2", "D1:D2", "E1:G1", "H1:J1"]:
        ws.merge_cells(rng)

    center = Alignment(horizontal="center", vertical="center")
    bold = Font(bold=True)
    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=10):
        for cell in row:
            cell.alignment = center
            cell.font = bold

    r = 3
    for item in per_slice_rows:
        ws.cell(r, 1, item["patient_id"])
        ws.cell(r, 2, item["slice_id"])
        ws.cell(r, 3, item["lower"])
        ws.cell(r, 4, item["upper"])
        ws.cell(r, 5, item["dice_zero"])
        ws.cell(r, 6, item["dice_train"])
        ws.cell(r, 7, item["dice_nnunet"])
        ws.cell(r, 8, item["hd95_zero"])
        ws.cell(r, 9, item["hd95_train"])
        ws.cell(r, 10, item["hd95_nnunet"])
        r += 1

    for rr in range(3, r):
        for cc in range(1, 11):
            ws.cell(rr, cc).alignment = center
    for rr in range(3, r):
        for cc in [5, 6, 7, 8, 9, 10]:
            ws.cell(rr, cc).number_format = "0.00"

    for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]:
        ws.column_dimensions[col].width = 14


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--gt-root",
        type=Path,
        default=Path("/home/wusi/SAM2/SAM2data/Eso/20260326/datanii/test_nii"),
    )
    parser.add_argument(
        "--train-pred-root",
        type=Path,
        default=Path("/home/wusi/SAM2/SAM2data/Eso/20260326/two_epoch/oracle_mask/mask_prompt_2/TestResult/best_mask"),
    )
    parser.add_argument(
        "--zero-pred-root",
        type=Path,
        default=Path("/home/wusi/SAM2/SAM2data/Eso/20260326/zero-shot/oracle_mask/mask_prompt_2/best_mask"),
    )
    parser.add_argument(
        "--nnunet-pred-root",
        type=Path,
        default=Path(
            "/home/wusi/nnUNet/nnUNetFrame/DATASET/nnUNet_results/"
            "Dataset008_EsoCTV73p/nnUNetTrainer__nnUNetPlans__3d_fullres/testResult83p_fold1"
        ),
    )
    parser.add_argument(
        "--table-path",
        type=Path,
        default=Path("/home/wusi/SAM2/SAM2data/Eso/20260326/prompt_layer_merge_train_zeroshot.xlsx"),
    )
    parser.add_argument("--dice-check-tol", type=float, default=1e-3)
    args = parser.parse_args()

    if not args.gt_root.exists():
        raise FileNotFoundError(args.gt_root)
    if not args.train_pred_root.exists():
        raise FileNotFoundError(args.train_pred_root)
    if not args.zero_pred_root.exists():
        raise FileNotFoundError(args.zero_pred_root)
    if not args.nnunet_pred_root.exists():
        raise FileNotFoundError(args.nnunet_pred_root)
    if not args.table_path.exists():
        raise FileNotFoundError(args.table_path)

    cases = collect_case_paths(
        args.gt_root,
        args.train_pred_root,
        args.zero_pred_root,
        args.nnunet_pred_root,
    )
    if len(cases) == 0:
        raise RuntimeError("No valid cases found to evaluate.")

    # Evaluate 3D + per-slice 2D
    results_3d: Dict[str, Dict[str, float]] = {}
    per_slice_rows: List[Dict] = []

    for case in cases:
        gt_zyx, gt_img = read_nii(case.gt_path)
        tr_zyx, tr_img = read_nii(case.train_pred_path)
        zs_zyx, zs_img = read_nii(case.zero_pred_path)
        nn_zyx, nn_img = read_nii(case.nnunet_pred_path)

        gt = to_binary(gt_zyx)
        tr = to_binary(tr_zyx)
        zs = to_binary(zs_zyx)
        nn = to_binary(nn_zyx)

        if gt.shape != tr.shape or gt.shape != zs.shape or gt.shape != nn.shape:
            raise ValueError(
                f"Shape mismatch for {case.patient_id}: "
                f"gt{gt.shape}, train{tr.shape}, zero{zs.shape}, nnunet{nn.shape}"
            )

        spacing_xyz = tuple(float(x) for x in gt_img.GetSpacing())  # (x,y,z)
        lower, upper, pos_slices = get_gt_bounds(gt)
        if lower < 0:
            print(f"[WARN] {case.patient_id}: GT empty, skipped")
            continue

        dice_train = dice_nd(tr, gt)
        dice_zero = dice_nd(zs, gt)
        dice_nnunet = dice_nd(nn, gt)
        hd95_train = hd95_binary(tr, gt, spacing_xyz)
        hd95_zero = hd95_binary(zs, gt, spacing_xyz)
        hd95_nnunet = hd95_binary(nn, gt, spacing_xyz)

        results_3d[case.patient_id] = {
            "lower": lower,
            "upper": upper,
            "dice_zero": dice_zero,
            "dice_train": dice_train,
            "dice_nnunet": dice_nnunet,
            "hd95_zero": hd95_zero,
            "hd95_train": hd95_train,
            "hd95_nnunet": hd95_nnunet,
        }

        # 2D metrics on GT-positive slices only
        spacing_xy = (spacing_xyz[0], spacing_xyz[1])  # (x,y)
        for z in pos_slices:
            gt2 = gt[z]
            tr2 = tr[z]
            zs2 = zs[z]
            nn2 = nn[z]
            per_slice_rows.append(
                {
                    "patient_id": case.patient_id,
                    "slice_id": z,
                    "lower": lower,
                    "upper": upper,
                    "dice_zero": dice_nd(zs2, gt2),
                    "dice_train": dice_nd(tr2, gt2),
                    "dice_nnunet": dice_nd(nn2, gt2),
                    "hd95_zero": hd95_binary(zs2, gt2, spacing_xy),
                    "hd95_train": hd95_binary(tr2, gt2, spacing_xy),
                    "hd95_nnunet": hd95_binary(nn2, gt2, spacing_xy),
                }
            )

    # Load workbook and compare old Dice in first sheet
    wb = load_workbook(args.table_path)
    ws = wb[wb.sheetnames[0]]

    # Existing first-sheet row map (assume data starts row 3)
    row_map: Dict[str, int] = {}
    old_vals: Dict[str, Tuple[Optional[float], Optional[float]]] = {}
    r = 3
    while True:
        pid = ws.cell(r, 1).value
        if pid is None or str(pid).strip() == "":
            break
        pid = str(pid).strip()
        row_map[pid] = r
        old_zero = cell_float(ws.cell(r, 4).value)
        old_train = cell_float(ws.cell(r, 5).value)
        old_vals[pid] = (old_zero, old_train)
        r += 1

    mismatches: List[Tuple[str, Optional[float], Optional[float], Optional[float], Optional[float]]] = []
    for pid, d in results_3d.items():
        old_zero, old_train = old_vals.get(pid, (None, None))
        if old_zero is not None and abs(old_zero - d["dice_zero"]) > args.dice_check_tol:
            mismatches.append((pid, old_zero, d["dice_zero"], old_train, d["dice_train"]))
        elif old_train is not None and abs(old_train - d["dice_train"]) > args.dice_check_tol:
            mismatches.append((pid, old_zero, d["dice_zero"], old_train, d["dice_train"]))

    # Overwrite first sheet with current results (+HD95_3D)
    write_main_sheet(ws, results_3d, mismatches)
    # Write second sheet per-slice
    write_slice_sheet(wb, per_slice_rows)

    wb.save(args.table_path)
    print(f"[DONE] Updated table: {args.table_path}")
    print(f"[DONE] 3D cases written: {len(results_3d)}")
    print(f"[DONE] 2D rows written: {len(per_slice_rows)}")


if __name__ == "__main__":
    main()

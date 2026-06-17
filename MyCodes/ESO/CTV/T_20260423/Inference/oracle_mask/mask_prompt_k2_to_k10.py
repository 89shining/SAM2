#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
SAM2 oracle inference using pre-defined prompt slices from Excel (K=2..10).

Workflow:
1) Load prompt slice lists from Oracle_Summary.xlsx sheets K2..K10.
2) Keep only patients p_55..p_82.
3) For each patient and each K, run one-shot inference with all K prompts at once.
4) Save predictions into separate folders per K:
      <OUT_ROOT>/K2/CTV_055.nii.gz
      ...
      <OUT_ROOT>/K10/CTV_082.nii.gz
"""

import ast
import os
import re
import shutil
import sys
import tempfile
from pathlib import Path

import numpy as np
import SimpleITK as sitk
import torch
from openpyxl import load_workbook
from PIL import Image

# Keep compatible with existing environment settings.
sys.path.append("/home/wusi/SAM2")
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from sam2.build_sam import build_sam2_video_predictor


# ======================================================
# ================== Path & Config =====================
# ======================================================

DATA_ROOT = Path("/home/wusi/SAM2/SAM2data/Eso/20260108/Data83")
OUT_ROOT = Path("/home/wusi/SAM2/SAM2data/Eso/20260423/Zero-shot/oracle_mask/mask_prompt")

# You can override via env ORACLE_XLSX
ORACLE_XLSX = Path(
    os.environ.get(
        "ORACLE_XLSX",
        r"/home/wusi/SAM2/SAM2data/Eso/20260108/Statistics/AAPM/Oracle_Summary.xlsx",
    )
)

IMG_NAME = "image.nii.gz"
GT_NAME = "CTV.nii.gz"

SAM2_CKPT = Path("/home/wusi/SAM2/checkpoints/sam2.1_hiera_large.pt")
SAM2_CFG = "configs/sam2.1/sam2.1_hiera_l.yaml"

DEVICE = "cuda"  # "cuda" or "cpu"
OBJ_ID = 1

K_VALUES = list(range(2, 11))
PID_MIN = 55
PID_MAX = 82

# CT window
WINDOW_CENTER = 40
WINDOW_WIDTH = 400


# ======================================================
# ======================= Utils =========================
# ======================================================


def window_to_uint8(img2d, wc, ww):
    img = img2d.astype(np.float32)
    lo = wc - ww / 2.0
    hi = wc + ww / 2.0
    img = np.clip(img, lo, hi)
    img = (img - lo) / (hi - lo + 1e-6) * 255.0
    return img.astype(np.uint8)


def save_frames_from_volume(vol_zyx, out_dir):
    out_dir.mkdir(parents=True, exist_ok=True)
    for i in range(vol_zyx.shape[0]):
        u8 = window_to_uint8(vol_zyx[i], WINDOW_CENTER, WINDOW_WIDTH)
        rgb = np.stack([u8, u8, u8], axis=-1)
        Image.fromarray(rgb).save(out_dir / f"{i:05d}.jpg", quality=95)


def read_nii_zyx(path):
    img = sitk.ReadImage(str(path))
    arr = sitk.GetArrayFromImage(img)
    return arr, img


def write_mask_like(pred_zyx, ref_img, out_path):
    pred_zyx = (pred_zyx > 0).astype(np.uint8)
    out = sitk.GetImageFromArray(pred_zyx)
    out.SetSpacing(ref_img.GetSpacing())
    out.SetOrigin(ref_img.GetOrigin())
    out.SetDirection(ref_img.GetDirection())
    sitk.WriteImage(out, str(out_path))


def parse_prompt_slices(value):
    if value is None:
        return []

    if isinstance(value, (list, tuple)):
        return [int(v) for v in value]

    text = str(value).strip()
    if text == "":
        return []

    # Expected format: "[0, 28, 17]"
    try:
        parsed = ast.literal_eval(text)
        if isinstance(parsed, (list, tuple)):
            return [int(v) for v in parsed]
    except Exception:
        pass

    # Fallback: extract integers from arbitrary text
    nums = re.findall(r"-?\d+", text)
    return [int(v) for v in nums]


def load_prompt_table(xlsx_path):
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Oracle summary not found: {xlsx_path}")

    wb = load_workbook(str(xlsx_path), data_only=True)
    out = {k: {} for k in K_VALUES}

    for k in K_VALUES:
        sname = f"K{k}"
        if sname not in wb.sheetnames:
            raise RuntimeError(f"Missing sheet '{sname}' in {xlsx_path}")

        ws = wb[sname]
        header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        idx = {str(v).strip(): i + 1 for i, v in enumerate(header) if v is not None}

        needed = ["PatientID", "PromptSlices"]
        for key in needed:
            if key not in idx:
                raise RuntimeError(f"Sheet '{sname}' missing column '{key}'")

        col_pid = idx["PatientID"]
        col_slices = idx["PromptSlices"]

        for r in range(2, ws.max_row + 1):
            pid = ws.cell(r, col_pid).value
            slc = ws.cell(r, col_slices).value
            if pid is None:
                continue

            pid_text = str(pid).strip()
            m = re.fullmatch(r"p_(\d+)", pid_text)
            if m is None:
                continue

            pid_num = int(m.group(1))
            if not (PID_MIN <= pid_num <= PID_MAX):
                continue

            out[k][pid_text] = parse_prompt_slices(slc)

    return out


# ======================================================
# =================== SAM2 Inference ====================
# ======================================================


@torch.no_grad()
def infer_with_prompt_list(predictor, frame_dir, gt_zyx, prompt_ids):
    state = predictor.init_state(video_path=str(frame_dir))
    predictor.reset_state(state)

    # Keep order and remove accidental duplicates
    prompt_ids = list(dict.fromkeys(int(x) for x in prompt_ids))

    z = gt_zyx.shape[0]
    for sid in prompt_ids:
        if sid < 0 or sid >= z:
            raise RuntimeError(f"Prompt slice out of range: {sid} (z={z})")

        prompt_mask = (gt_zyx[sid] > 0).astype(np.uint8)
        if prompt_mask.sum() == 0:
            raise RuntimeError(f"Prompt slice {sid} is empty in GT.")

        predictor.add_new_mask(
            inference_state=state,
            frame_idx=sid,
            obj_id=OBJ_ID,
            mask=prompt_mask,
        )

    _, h, w = gt_zyx.shape
    pred = np.zeros((z, h, w), dtype=np.uint8)

    for fidx, obj_ids, logits in predictor.propagate_in_video(state):
        for i, oid in enumerate(obj_ids):
            if int(oid) == OBJ_ID:
                pred[int(fidx)] = (logits[i] > 0).cpu().numpy()
                break

    return pred


# ======================================================
# ======================== Main =========================
# ======================================================


def main():
    OUT_ROOT.mkdir(parents=True, exist_ok=True)
    for k in K_VALUES:
        (OUT_ROOT / f"K{k}").mkdir(parents=True, exist_ok=True)

    prompt_table = load_prompt_table(ORACLE_XLSX)
    print(f"[INFO] Loaded prompt table: {ORACLE_XLSX}")

    device = torch.device(
        DEVICE if (DEVICE == "cpu" or torch.cuda.is_available()) else "cpu"
    )
    predictor = build_sam2_video_predictor(SAM2_CFG, str(SAM2_CKPT), device=device)

    patient_dirs = []
    for p in sorted(DATA_ROOT.iterdir()):
        if not p.is_dir():
            continue
        m = re.fullmatch(r"p_(\d+)", p.name)
        if m is None:
            continue
        pid_num = int(m.group(1))
        if PID_MIN <= pid_num <= PID_MAX:
            patient_dirs.append(p)

    print(f"[INFO] Target patients in data root: {len(patient_dirs)} (p_{PID_MIN}..p_{PID_MAX})")

    for pdir in patient_dirs:
        pid_num = int(re.fullmatch(r"p_(\d+)", pdir.name).group(1))
        patient_id = f"CTV_{pid_num:03d}"

        img_path = pdir / IMG_NAME
        gt_path = pdir / GT_NAME
        if not img_path.exists() or not gt_path.exists():
            print(f"[WARN] Skip {pdir.name}: missing image or GT")
            continue

        img_zyx, img_sitk = read_nii_zyx(img_path)
        gt_zyx, _ = read_nii_zyx(gt_path)
        gt_zyx = (gt_zyx > 0).astype(np.uint8)

        if img_zyx.shape != gt_zyx.shape:
            print(
                f"[WARN] Skip {pdir.name}: shape mismatch img{img_zyx.shape} vs gt{gt_zyx.shape}"
            )
            continue

        tmp_dir = Path(tempfile.mkdtemp(prefix=f"sam2_{pdir.name}_"))
        try:
            save_frames_from_volume(img_zyx, tmp_dir)

            for k in K_VALUES:
                prompts = prompt_table.get(k, {}).get(pdir.name)
                if prompts is None:
                    print(f"[WARN] {pdir.name} missing PromptSlices in sheet K{k}, skip")
                    continue

                if len(prompts) == 0:
                    print(f"[WARN] {pdir.name} K{k} prompt list empty, skip")
                    continue

                try:
                    pred = infer_with_prompt_list(
                        predictor=predictor,
                        frame_dir=tmp_dir,
                        gt_zyx=gt_zyx,
                        prompt_ids=prompts,
                    )
                except Exception as e:
                    print(f"[WARN] {pdir.name} K{k} inference failed: {e}")
                    continue

                out_mask_path = OUT_ROOT / f"K{k}" / f"{patient_id}.nii.gz"
                write_mask_like(pred, img_sitk, out_mask_path)
                print(f"[OK] {pdir.name} K{k} -> {out_mask_path} | prompts={prompts}")

        finally:
            shutil.rmtree(tmp_dir, ignore_errors=True)

    print("[DONE] All requested K inference finished.")


if __name__ == "__main__":
    main()

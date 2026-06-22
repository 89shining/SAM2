#!/usr/bin/env python
# -*- coding: utf-8 -*-

from __future__ import annotations

import argparse
import os
import shutil
import subprocess
import sys
from pathlib import Path

from io_utils import DEFAULT_INIT_CKPT, DEFAULT_MODEL_CFG


CURRENT_DIR = Path(__file__).resolve().parent
GTVP_DATA_ROOT = Path("/home/intern/ftp/wusi/SAM2/MyTrain/SAM2data/Rectal/20260616_GTVp/datanii")
GTVP_ALIAS_ROOT = GTVP_DATA_ROOT.parent / "datanii_ctv_alias"


def run_cmd(cmd: list, gpu_id: str | None = None):
    print("[RUN]", " ".join(str(x) for x in cmd), flush=True)
    env = os.environ.copy()
    if gpu_id is not None and str(gpu_id).strip():
        env["CUDA_VISIBLE_DEVICES"] = str(gpu_id).strip()
        print(f"[GPU] CUDA_VISIBLE_DEVICES={env['CUDA_VISIBLE_DEVICES']}", flush=True)
    subprocess.run([str(x) for x in cmd], check=True, env=env)


def _patient_sort_key(path: Path):
    import re

    parts = re.split(r"(\d+)", path.name)
    return [int(x) if x.isdigit() else x.lower() for x in parts]


def _link_or_copy(src: Path, dst: Path):
    if dst.exists() or dst.is_symlink():
        return
    try:
        dst.symlink_to(src)
    except OSError:
        shutil.copy2(src, dst)


def prepare_ctv_alias_dataset(src_root: Path, alias_root: Path):
    for split in ("train_nii", "test_nii"):
        src_split = src_root / split
        if not src_split.exists():
            raise FileNotFoundError(f"Missing split directory: {src_split}")
        for src_patient in sorted([p for p in src_split.iterdir() if p.is_dir()], key=_patient_sort_key):
            image_src = src_patient / "image.nii.gz"
            mask_src = src_patient / "GTVp.nii.gz"
            if not image_src.exists() or not mask_src.exists():
                raise FileNotFoundError(f"Missing image.nii.gz or GTVp.nii.gz in {src_patient}")

            dst_patient = alias_root / split / src_patient.name
            dst_patient.mkdir(parents=True, exist_ok=True)
            _link_or_copy(image_src, dst_patient / "image.nii.gz")
            _link_or_copy(mask_src, dst_patient / "CTV.nii.gz")


def rename_outputs_to_gtvp(result_root: Path):
    test_root = result_root / "TestResults"
    if not test_root.exists():
        return
    for nii_path in test_root.glob("Model_*/Prompt_*/CTV_*.nii.gz"):
        new_name = nii_path.name.replace("CTV_", "GTVp_", 1)
        new_path = nii_path.with_name(new_name)
        if not new_path.exists():
            nii_path.rename(new_path)

    try:
        from openpyxl import load_workbook
    except Exception:
        return

    for xlsx_path in list(test_root.glob("Model_*/Model_*_evaluation.xlsx")) + [
        test_root / "test_prompt_frames.xlsx"
    ]:
        if not xlsx_path.exists():
            continue
        wb = load_workbook(xlsx_path)
        changed = False
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and "CTV_" in cell.value:
                        cell.value = cell.value.replace("CTV_", "GTVp_")
                        changed = True
        if changed:
            wb.save(xlsx_path)


def main():
    parser = argparse.ArgumentParser("Run GTVp prompt-number experiment")
    parser.add_argument("--gpu-id", type=str, default=None)
    parser.add_argument("--init-ckpt", type=Path, default=DEFAULT_INIT_CKPT)
    parser.add_argument("--model-cfg", type=str, default=DEFAULT_MODEL_CFG)
    parser.add_argument("--train-models", type=str, default="4")
    parser.add_argument("--test-models", type=str, default="2-4")
    parser.add_argument("--test-ks", type=str, default="2-4")
    parser.add_argument("--folds", type=str, default="0-4")
    parser.add_argument("--seed", type=int, default=20260616)
    parser.add_argument("--max-epochs", type=int, default=100)
    parser.add_argument("--patience", type=int, default=15)
    parser.add_argument("--lr", type=float, default=1e-4)
    parser.add_argument("--min-lr", type=float, default=1e-6)
    parser.add_argument("--warmup-epochs", type=int, default=5)
    parser.add_argument("--weight-decay", type=float, default=1e-4)
    parser.add_argument("--grad-clip-norm", type=float, default=1.0)
    parser.add_argument("--input-size", type=int, default=512)
    parser.add_argument("--window-center", type=float, default=40.0)
    parser.add_argument("--window-width", type=float, default=400.0)
    parser.add_argument("--num-workers", type=int, default=2)
    parser.add_argument("--device", type=str, default="cuda")
    parser.add_argument("--amp-dtype", type=str, default="bfloat16", choices=["bfloat16", "float16"])
    parser.add_argument("--lora-r", type=int, default=4)
    parser.add_argument("--lora-alpha", type=int, default=16)
    parser.add_argument("--lora-dropout", type=float, default=0.1)
    parser.add_argument("--bidirectional-train", action="store_true")
    parser.add_argument("--forward-backbone-per-frame", action="store_true")
    parser.add_argument("--no-train", action="store_true")
    parser.add_argument("--no-test", action="store_true")
    parser.add_argument("--no-resume", action="store_true")
    parser.add_argument("--no-save-predictions", action="store_true")
    args = parser.parse_args()

    prepare_ctv_alias_dataset(GTVP_DATA_ROOT, GTVP_ALIAS_ROOT)

    common_args = [
        "--data-root",
        GTVP_ALIAS_ROOT,
        "--init-ckpt",
        args.init_ckpt,
        "--model-cfg",
        args.model_cfg,
        "--seed",
        args.seed,
        "--input-size",
        args.input_size,
        "--window-center",
        args.window_center,
        "--window-width",
        args.window_width,
        "--num-workers",
        args.num_workers,
        "--device",
        args.device,
        "--amp-dtype",
        args.amp_dtype,
        "--lora-r",
        args.lora_r,
        "--lora-alpha",
        args.lora_alpha,
        "--lora-dropout",
        args.lora_dropout,
    ]

    if not args.no_train:
        train_cmd = [
            sys.executable,
            CURRENT_DIR / "train.py",
            *common_args,
            "--models",
            args.train_models,
            "--max-epochs",
            args.max_epochs,
            "--patience",
            args.patience,
            "--lr",
            args.lr,
            "--min-lr",
            args.min_lr,
            "--warmup-epochs",
            args.warmup_epochs,
            "--weight-decay",
            args.weight_decay,
            "--grad-clip-norm",
            args.grad_clip_norm,
        ]
        if args.bidirectional_train:
            train_cmd.append("--bidirectional-train")
        if args.forward_backbone_per_frame:
            train_cmd.append("--forward-backbone-per-frame")
        if args.no_resume:
            train_cmd.append("--no-resume")
        run_cmd(train_cmd, gpu_id=args.gpu_id)

    if not args.no_test:
        test_cmd = [
            sys.executable,
            CURRENT_DIR / "test.py",
            *common_args,
            "--models",
            args.test_models,
            "--test-ks",
            args.test_ks,
            "--folds",
            args.folds,
        ]
        if args.forward_backbone_per_frame:
            test_cmd.append("--forward-backbone-per-frame")
        if args.no_save_predictions:
            test_cmd.append("--no-save-predictions")
        run_cmd(test_cmd, gpu_id=args.gpu_id)
        rename_outputs_to_gtvp(GTVP_DATA_ROOT.parent)


if __name__ == "__main__":
    main()

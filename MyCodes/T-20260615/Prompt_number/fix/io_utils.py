#!/usr/bin/env python
# -*- coding: utf-8 -*-

from __future__ import annotations

import json
import os
import random
import re
import sys
from pathlib import Path
from typing import Sequence

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import SimpleITK as sitk
import torch
import torch.nn.functional as F
from openpyxl import Workbook
from hydra import compose, initialize_config_module
from hydra.core.global_hydra import GlobalHydra
from hydra.utils import instantiate
from omegaconf import OmegaConf
from torch.utils.data import Dataset

CURRENT_DIR = Path(__file__).resolve().parent


def find_project_root(start: Path = CURRENT_DIR) -> Path:
    for p in [start] + list(start.parents):
        if (p / "training").is_dir() and (p / "sam2").is_dir():
            return p
    env_root = os.environ.get("SAM2_PROJECT_ROOT", "").strip()
    if env_root:
        p = Path(env_root).resolve()
        if (p / "training").is_dir() and (p / "sam2").is_dir():
            return p
    raise RuntimeError("Cannot locate SAM2 project root. Set SAM2_PROJECT_ROOT if needed.")


PROJECT_ROOT = find_project_root()
if str(CURRENT_DIR) not in sys.path:
    sys.path.insert(0, str(CURRENT_DIR))
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from training.model.sam2 import SAM2Train
from training.utils.data_utils import Frame, Object, VideoDatapoint
from experiment_core import configure_prompt_number_trainables


DEFAULT_DATA_ROOT = Path("/home/intern/ftp/wusi/SAM2/MyTrain/SAM2data/Rectal/20260616_CTV/datanii")
DEFAULT_MODEL_CFG = "configs/sam2.1/sam2.1_hiera_s.yaml"
DEFAULT_INIT_CKPT = Path("/home/intern/ftp/wusi/SAM2/checkpoints/sam2.1_hiera_small.pt")


def default_result_base(data_root: Path) -> Path:
    return Path(data_root).parent / "Prompt_number" / "fix"


def set_global_seed(seed: int):
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)


def patient_sort_key(path_obj: Path):
    parts = re.split(r"(\d+)", path_obj.name)
    return [int(x) if x.isdigit() else x.lower() for x in parts]


def patient_id_from_dir(pdir: Path) -> int:
    m = re.search(r"(\d+)$", pdir.name)
    if m is None:
        raise ValueError(f"Cannot parse patient id from {pdir}")
    return int(m.group(1))


def ctv_case_name(pdir: Path) -> str:
    return f"CTV_{patient_id_from_dir(pdir):03d}"


def window_to_uint8(img2d: np.ndarray, wc: float, ww: float) -> np.ndarray:
    img = img2d.astype(np.float32)
    lo = wc - ww / 2.0
    hi = wc + ww / 2.0
    img = np.clip(img, lo, hi)
    return ((img - lo) / (hi - lo + 1e-6) * 255.0).astype(np.uint8)


class RectalCTVVolumeDataset(Dataset):
    def __init__(
        self,
        patient_dirs: Sequence[Path],
        image_name: str = "image.nii.gz",
        mask_name: str = "CTV.nii.gz",
        input_size: int = 512,
        window_center: float = 40.0,
        window_width: float = 400.0,
    ):
        self.patient_dirs = [Path(p) for p in patient_dirs]
        self.image_name = image_name
        self.mask_name = mask_name
        self.input_size = int(input_size)
        self.window_center = float(window_center)
        self.window_width = float(window_width)
        self.samples = []
        for pdir in self.patient_dirs:
            if (pdir / image_name).exists() and (pdir / mask_name).exists():
                self.samples.append(pdir)

    def __len__(self):
        return len(self.samples)

    def __getitem__(self, idx):
        pdir = self.samples[idx]
        img = sitk.GetArrayFromImage(sitk.ReadImage(str(pdir / self.image_name)))
        gt = sitk.GetArrayFromImage(sitk.ReadImage(str(pdir / self.mask_name)))
        gt = (gt > 0).astype(np.uint8)
        if img.shape != gt.shape:
            raise ValueError(f"Shape mismatch in {pdir}: image {img.shape}, mask {gt.shape}")

        frames = []
        h0, w0 = img.shape[1], img.shape[2]
        for t in range(img.shape[0]):
            u8 = window_to_uint8(img[t], self.window_center, self.window_width)
            rgb = np.stack([u8, u8, u8], axis=0)
            image_tensor = torch.from_numpy(rgb).float() / 255.0
            image_tensor = F.interpolate(
                image_tensor.unsqueeze(0),
                size=(self.input_size, self.input_size),
                mode="bilinear",
                align_corners=False,
            ).squeeze(0)

            mask_tensor = torch.from_numpy(gt[t]).float().unsqueeze(0).unsqueeze(0)
            mask_tensor = F.interpolate(mask_tensor, size=(self.input_size, self.input_size), mode="nearest")
            mask_tensor = mask_tensor.squeeze(0).squeeze(0).to(torch.bool)
            frames.append(Frame(data=image_tensor, objects=[Object(object_id=1, frame_index=t, segment=mask_tensor)]))

        return VideoDatapoint(frames=frames, video_id=patient_id_from_dir(pdir), size=(h0, w0))


def list_patient_dirs(root: Path) -> list[Path]:
    if not root.exists():
        raise FileNotFoundError(f"Patient root not found: {root}")
    return sorted([p for p in root.iterdir() if p.is_dir()], key=patient_sort_key)


def make_or_load_splits(patient_dirs: Sequence[Path], num_folds: int, seed: int, split_path: Path) -> list[dict]:
    patient_dirs = list(patient_dirs)
    id_to_path = {p.name: p for p in patient_dirs}
    if split_path.exists():
        data = json.loads(split_path.read_text(encoding="utf-8"))
        return [
            {
                "fold": item["fold"],
                "train": [id_to_path[x] for x in item["train"]],
                "val": [id_to_path[x] for x in item["val"]],
            }
            for item in data["folds"]
        ]

    rng = np.random.RandomState(seed)
    idx = np.arange(len(patient_dirs))
    rng.shuffle(idx)
    chunks = np.array_split(idx, num_folds)
    folds = []
    for fold in range(num_folds):
        val_idx = set(int(x) for x in chunks[fold].tolist())
        train = [patient_dirs[i] for i in range(len(patient_dirs)) if i not in val_idx]
        val = [patient_dirs[i] for i in range(len(patient_dirs)) if i in val_idx]
        folds.append({"fold": fold, "train": train, "val": val})

    split_path.parent.mkdir(parents=True, exist_ok=True)
    serializable = {
        "seed": seed,
        "num_folds": num_folds,
        "folds": [
            {"fold": f["fold"], "train": [p.name for p in f["train"]], "val": [p.name for p in f["val"]]}
            for f in folds
        ],
    }
    split_path.write_text(json.dumps(serializable, indent=2), encoding="utf-8")
    return folds


def load_model_cfg_dict(model_cfg: str):
    if GlobalHydra.instance().is_initialized():
        GlobalHydra.instance().clear()
    with initialize_config_module("sam2", version_base="1.2"):
        cfg = compose(config_name=model_cfg)
    return OmegaConf.to_container(cfg.model, resolve=True)


def build_model(model_cfg: str, init_ckpt: Path, device: torch.device, args):
    model_cfg_dict = load_model_cfg_dict(model_cfg)
    model_cfg_dict["image_size"] = int(args.input_size)
    model_cfg_dict["freeze_image_encoder"] = False
    image_encoder_cfg = model_cfg_dict.pop("image_encoder")
    memory_attention_cfg = model_cfg_dict.pop("memory_attention")
    memory_encoder_cfg = model_cfg_dict.pop("memory_encoder")
    model_cfg_dict.pop("_target_", None)

    model = SAM2Train(
        image_encoder=instantiate(image_encoder_cfg, _recursive_=True),
        memory_attention=instantiate(memory_attention_cfg, _recursive_=True),
        memory_encoder=instantiate(memory_encoder_cfg, _recursive_=True),
        prob_to_use_pt_input_for_train=0.0,
        prob_to_use_pt_input_for_eval=0.0,
        prob_to_use_box_input_for_train=0.0,
        prob_to_use_box_input_for_eval=0.0,
        prob_to_sample_from_gt_for_train=0.0,
        num_frames_to_correct_for_train=1,
        num_frames_to_correct_for_eval=1,
        rand_frames_to_correct_for_train=False,
        rand_frames_to_correct_for_eval=False,
        add_all_frames_to_correct_as_cond=False,
        num_correction_pt_per_frame=0,
        rand_init_cond_frames_for_train=False,
        rand_init_cond_frames_for_eval=False,
        **model_cfg_dict,
    )

    state = torch.load(str(init_ckpt), map_location="cpu")
    state_dict = state["model"] if isinstance(state, dict) and "model" in state else state
    model.load_state_dict(state_dict, strict=False)
    stats = configure_prompt_number_trainables(
        model,
        lora_r=args.lora_r,
        lora_alpha=args.lora_alpha,
        lora_dropout=args.lora_dropout,
    )
    return model.to(device), stats


def build_optimizer(model, lr: float, weight_decay: float):
    params = [p for p in model.parameters() if p.requires_grad]
    return torch.optim.AdamW(params, lr=float(lr), weight_decay=float(weight_decay))


def build_scheduler(optimizer, max_epochs: int, warmup_epochs: int, min_lr: float):
    base_lr = float(optimizer.param_groups[0]["lr"])
    min_factor = float(min_lr) / max(base_lr, 1e-12)

    def lr_lambda(epoch):
        if warmup_epochs > 0 and epoch < warmup_epochs:
            return float(epoch + 1) / float(warmup_epochs)
        progress = (epoch - warmup_epochs) / max(1, max_epochs - warmup_epochs)
        cosine = 0.5 * (1.0 + np.cos(np.pi * progress))
        return max(min_factor, float(cosine))

    return torch.optim.lr_scheduler.LambdaLR(optimizer, lr_lambda=lr_lambda)


def save_checkpoint(path: Path, epoch: int, model, optimizer, scheduler, scaler, best_metric: float, best_epoch: int, args_dict: dict):
    path.parent.mkdir(parents=True, exist_ok=True)
    torch.save(
        {
            "epoch": int(epoch),
            "model": model.state_dict(),
            "optimizer": optimizer.state_dict(),
            "scheduler": scheduler.state_dict() if scheduler else None,
            "scaler": scaler.state_dict() if scaler else None,
            "best_metric": float(best_metric),
            "best_epoch": int(best_epoch),
            "args": args_dict,
        },
        str(path),
    )


def load_checkpoint(path: Path, model, optimizer=None, scheduler=None, scaler=None, device="cpu"):
    state = torch.load(str(path), map_location=device)
    model.load_state_dict(state["model"], strict=False)
    if optimizer is not None and state.get("optimizer") is not None:
        optimizer.load_state_dict(state["optimizer"])
    if scheduler is not None and state.get("scheduler") is not None:
        scheduler.load_state_dict(state["scheduler"])
    if scaler is not None and state.get("scaler") is not None:
        scaler.load_state_dict(state["scaler"])
    return int(state.get("epoch", 0)), float(state.get("best_metric", -1.0)), int(state.get("best_epoch", -1))


def append_table_txt(path: Path, row: dict):
    path.parent.mkdir(parents=True, exist_ok=True)
    write_header = not path.exists()
    with open(path, "a", encoding="utf-8") as f:
        if write_header:
            f.write("\t".join(row.keys()) + "\n")
        f.write("\t".join(str(v) for v in row.values()) + "\n")


def append_txt_log(path: Path, text: str):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "a", encoding="utf-8") as f:
        f.write(text.rstrip() + "\n")


def plot_loss_curve(log_txt: Path, out_png: Path):
    if not log_txt.exists():
        return
    epochs, train_loss, val_metric = [], [], []
    with open(log_txt, "r", encoding="utf-8") as f:
        header = None
        for line in f:
            line = line.strip()
            if not line or line.startswith("=") or line.startswith("[") or ":" in line and "\t" not in line:
                continue
            parts = line.split("\t")
            if header is None:
                if "epoch" in parts and "train_loss" in parts:
                    header = parts
                continue
            if len(parts) != len(header):
                continue
            row = dict(zip(header, parts))
            if "epoch" not in row or "train_loss" not in row:
                continue
            epochs.append(int(row["epoch"]))
            train_loss.append(float(row["train_loss"]))
            val_metric.append(float(row["val_unprompted_slice_3d_dsc_mean"]))
    if not epochs:
        return
    out_png.parent.mkdir(parents=True, exist_ok=True)
    plt.figure(figsize=(8, 5))
    plt.plot(epochs, train_loss, label="train loss")
    plt.plot(epochs, val_metric, label="val unprompted-slice 3D DSC")
    plt.xlabel("epoch")
    plt.grid(True, alpha=0.3)
    plt.legend()
    plt.tight_layout()
    plt.savefig(out_png, dpi=160)
    plt.close()


def write_prompt_record_excel(
    path: Path,
    patient_dirs: Sequence[Path],
    prompt_ks: Sequence[int],
    split_name: str,
):
    from experiment_core import uniform_prompt_indices

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    for k in prompt_ks:
        ws = wb.create_sheet(f"Prompt_{k}")
        ws.append(["split", "patient", "num_frames", "prompt_k", "prompt_frame_ids"])
        for pdir in patient_dirs:
            mask_path = pdir / "CTV.nii.gz"
            if not mask_path.exists():
                continue
            num_frames = int(sitk.GetArrayFromImage(sitk.ReadImage(str(mask_path))).shape[0])
            prompts = uniform_prompt_indices(num_frames, int(k))
            ws.append([split_name, ctv_case_name(pdir), num_frames, int(k), ",".join(str(x) for x in prompts)])
    wb.save(path)


def write_training_prompt_plan_excel(
    path: Path,
    folds: Sequence[dict],
    model_xs: Sequence[int],
    max_epochs: int,
    seed: int,
):
    from experiment_core import uniform_prompt_indices

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    sheets = {}
    for k in range(2, max(int(x) for x in model_xs) + 1):
        ws = wb.create_sheet(f"Prompt_{k}")
        ws.append(["sampled_k", "model_x", "fold", "epoch", "patient", "num_frames", "prompt_frame_ids"])
        sheets[k] = ws

    for model_x in model_xs:
        for fold_info in folds:
            fold = int(fold_info["fold"])
            for epoch in range(int(max_epochs)):
                for pdir in fold_info["train"]:
                    mask_path = pdir / "CTV.nii.gz"
                    if not mask_path.exists():
                        continue
                    num_frames = int(sitk.GetArrayFromImage(sitk.ReadImage(str(mask_path))).shape[0])
                    prompts = uniform_prompt_indices(num_frames, model_x)
                    sampled_k = len(prompts)
                    sheets[sampled_k].append(
                        [
                            sampled_k,
                            int(model_x),
                            fold,
                            epoch + 1,
                            ctv_case_name(pdir),
                            num_frames,
                            ",".join(str(x) for x in prompts),
                        ]
                    )
    wb.save(path)


def write_model_evaluation_excel(path: Path, rows: list[dict], prompt_ks: Sequence[int]):
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"
    summary_ws.append(
        [
            "prompt_k",
            "unprompted_dice_3d",
            "unprompted_hd95",
            "unprompted_asd",
            "whole_dice_3d",
            "whole_hd95",
            "whole_asd",
        ]
    )

    for k in prompt_ks:
        k_rows = [r for r in rows if int(r["test_k"]) == int(k)]
        if not k_rows:
            continue

        metric_cols = [
            "unprompted_dice_3d",
            "unprompted_hd95",
            "unprompted_asd",
            "whole_dice_3d",
            "whole_hd95",
            "whole_asd",
        ]
        means = {m: float(np.mean([float(r[m]) for r in k_rows])) for m in metric_cols}
        stds = {m: float(np.std([float(r[m]) for r in k_rows], ddof=1)) if len(k_rows) > 1 else 0.0 for m in metric_cols}

        summary_ws.append(
            [
                int(k),
                f"{means['unprompted_dice_3d']:.2f}±{stds['unprompted_dice_3d']:.2f}",
                f"{means['unprompted_hd95']:.2f}±{stds['unprompted_hd95']:.2f}",
                f"{means['unprompted_asd']:.2f}±{stds['unprompted_asd']:.2f}",
                f"{means['whole_dice_3d']:.2f}±{stds['whole_dice_3d']:.2f}",
                f"{means['whole_hd95']:.2f}±{stds['whole_hd95']:.2f}",
                f"{means['whole_asd']:.2f}±{stds['whole_asd']:.2f}",
            ]
        )

        ws = wb.create_sheet(f"Prompt_{k}")
        ws.append(
            [
                "patient",
                "prompt_frames",
                "unprompted_dice_3d",
                "unprompted_hd95",
                "unprompted_asd",
                "whole_dice_3d",
                "whole_hd95",
                "whole_asd",
            ]
        )
        for r in k_rows:
            ws.append(
                [
                    r["patient"],
                    r["prompt_frames"],
                    round(float(r["unprompted_dice_3d"]), 2),
                    round(float(r["unprompted_hd95"]), 2),
                    round(float(r["unprompted_asd"]), 2),
                    round(float(r["whole_dice_3d"]), 2),
                    round(float(r["whole_hd95"]), 2),
                    round(float(r["whole_asd"]), 2),
                ]
            )
        ws.append(
            [
                "Mean",
                "",
                round(means["unprompted_dice_3d"], 2),
                round(means["unprompted_hd95"], 2),
                round(means["unprompted_asd"], 2),
                round(means["whole_dice_3d"], 2),
                round(means["whole_hd95"], 2),
                round(means["whole_asd"], 2),
            ]
        )
        ws.append(
            [
                "Std",
                "",
                round(stds["unprompted_dice_3d"], 2),
                round(stds["unprompted_hd95"], 2),
                round(stds["unprompted_asd"], 2),
                round(stds["whole_dice_3d"], 2),
                round(stds["whole_hd95"], 2),
                round(stds["whole_asd"], 2),
            ]
        )

    wb.save(path)
